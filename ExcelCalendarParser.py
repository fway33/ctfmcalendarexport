# This python file will parse the calendar dump excel spreadsheet
# and produce a text output suitable for editing.
import datetime
import re

import numpy
import openpyxl

from DataStructs import LodgeTown, BadLocations, locations, location_exceptions, full_list_calendar_entries, dinner, \
    degrees


#
# Method to find out how many rows in the sheet are populated.
# This differs (potentially) from max rows
#
def get_populated_row_count(ws):
    number_of_rows = ws.max_row
    last_row_index_with_data = 0

    while True:
        if ws.cell(number_of_rows, 1).value is not None:
            last_row_index_with_data = number_of_rows
            break
        else:
            number_of_rows -= 1
    return number_of_rows


def export_calendar_data():
    print("Here is where we'd start to parse")

    # Open the workbook.  There are two worksheets, 'data' and 'calc', we want 'data'
    wb = openpyxl.load_workbook('CalDump1.xlsx')
    ws = wb['data']

    # Find out how many rows are populated -- this is different than max rows.
    rows = get_populated_row_count(ws)
    print("ROWS " + str(rows))

    # initialize a list of lists, initialized to -1's, to hold our data when transformed
    # also initialize a list to hold locations when there is some doubt
    for i in range(rows - 1):
        full_list_calendar_entries.insert(i, [-1, -1, -1, -1, -1])
        locations.insert(i, '')
    print(full_list_calendar_entries)
    print(locations)
    print(full_list_calendar_entries[rows-1])
    if not full_list_calendar_entries[rows - 1]:
        full_list_calendar_entries.pop(rows - 1)

    # Set the ending cell identifier...column "E" with the last populated row number
    end = 'E' + str(rows)
    table = numpy.array([[cell.value for cell in col] for col in ws['A2':end]])

    # The first value in the row is the lodge name.  Need to change that from 'Lodge 001 Hiram'
    # to "Hiram Lodge No. 1".
    modify_lodge(table)
    modify_event_title(table)
    modify_event_descr(table)
    modify_event_location(table)
    modify_date(table)

    print("\n=-=-=-=-=-=-\n")
    print(full_list_calendar_entries)
    print(degrees)
    return full_list_calendar_entries


def modify_lodge(table):
    # Create an array of the first value in every list in table, the lists of lists
    # of the spreadsheet values.  These are all lodge names.
    lodge_array = [item[0] for item in table]

    # Since we will use it many times, compile the regex for the lodge name as
    # written in the cell (i.e., "Lodge 001 Hiram") and capture the values.
    rx = re.compile(r'(Lodge) (\d{3}) ([a-zA-Z\-\'\. ]+)')

    # Loop through the array elements and if the regex matches, the modify the string
    # (i.e., from 'Lodge 001 Hiram' to 'Hiram Lodge No. 1')
    array_idx = 0
    print(full_list_calendar_entries[array_idx])
    for lodge in lodge_array:
        result = rx.match(lodge)
        if result:
            lodge_string = result.group(3) + " " + result.group(1) + " No. " + result.group(2).lstrip("0")
            full_list_calendar_entries[array_idx][0] = lodge_string
        array_idx += 1
    print(full_list_calendar_entries)


def modify_event_title(table):
    # Create an array of the second value in every list in table, the lists of lists
    # of the spreadsheet values.  These are all event descriptions.
    event_array = [item[1] for item in table]
    print(event_array)

    rx = re.compile(r'.*[Dd]egree.*')

    array_idx = 0
    for event in event_array:
        result = rx.match(event)
        if result is not None:
            #            print(event, end="|")
            degrees.append(array_idx)
        full_list_calendar_entries[array_idx][1] = event
        array_idx += 1
    print()
    print(degrees)

def modify_event_descr(table):
    descr_array = [item[2] for item in table]
    print(descr_array)

    rx = re.compile(r'.*[Dd]egree.*')
    ry = re.compile(r'.*[Dd]inner.*')

    array_idx = 0
    for descr in descr_array:
        if descr:
            result = rx.match(descr)
            if result:
                if array_idx not in degrees:
                    degrees.append(array_idx)
            result2 = ry.search(descr)
            if result2:
                dinner.append(array_idx)
        full_list_calendar_entries[array_idx][2] = descr
        array_idx += 1


def modify_event_location(table):
    print("******* IN modify_even_location *******")
    loc_array = [item[3] for item in table]
    print(loc_array)
    print("........................")
    array_idx = 0
    rw = re.compile(r'.*[,]* (.*), (CT USA).*')
    rx = re.compile(r'[\w ]*[,]? (.*), (CT \d{5})[,]*.*')
    ry = re.compile(r'(Lodge) (\d{3}) ([a-zA-Z\-\'\. ]+)')
    rz = re.compile(r'.*, (.*), (CT)[,]* .*')
    ra = re.compile(r'.*, (.*)')

    # loop through the locations.  See what they match, if anything
    for loc in loc_array:
        print("---------------------------------------------------")
        print(loc)
        if not loc:
            # there was no location specified in the spreadsheet, so leave it empty in the result
            print("NO LOC ------")
            print(table[array_idx][0])
            print(LodgeTown[table[array_idx][0]])
            full_list_calendar_entries[array_idx][3] = LodgeTown[table[array_idx][0]]
        else:
            print(loc)

            # start by trying to pull the town out.  BUT, this might be a legit address for a
            # non-lodge event (say pizza night at a pizza place).  Leave the location
            # cell alone, but put the possible town into a separate location array
            result = rx.match(loc)
            if result:
                print("Match rx")
                print(result.group(1))
                result2 = ra.match(result.group(1))
                if result2:
                    print(result2.group(1))
                    locations[array_idx] = result2.group(1)
                else:
                    locations[array_idx] = result.group(1)
                full_list_calendar_entries[array_idx][3] = loc
            else:
                # See if 'CT' is by itself.
                result2 = rz.match(loc)
                if result2 is not None:
                    print("Match rz")
                    print(result2.group(1))
                    locations[array_idx] = result2.group(1)
                    full_list_calendar_entries[array_idx][3] = loc
                else:
                    # It might be CT USA, which sometimes shows up
                    result2 = rw.match(loc)
                    if result2:
                        print("Match rw")
                        print(result2.group(1))
                        locations[array_idx] = result2.group(1)
                        full_list_calendar_entries[array_idx][3] = loc
                    else:
                        # it didn't have 'CT' in the location, so see if it is maybe a Lodge 001 Hiram
                        # situation.  If so, use the lodge/town dictionary to get the address.
                        # (could probably just use the dictionary when we grab the lodge but this
                        # is a little more flexible)
                        result2 = ry.fullmatch(loc)
                        if result2:
                            print("match ry -- fullmatch)")
                            print(result2)
                            print(LodgeTown[loc])
                            full_list_calendar_entries[array_idx][3] = LodgeTown[loc]
                        else:
                            result2 = ry.search(loc)
                            if result2:
                                print("Match ry -- search")
                                print(result2)
                                lodge = result2.group(1) + " " + result2.group(2) + " " + result2.group(3)
                                if lodge in LodgeTown.keys():
                                    print(lodge)
                                    print(LodgeTown[lodge])
                                    print(array_idx)
                                    print(full_list_calendar_entries[array_idx])
                                    full_list_calendar_entries[array_idx][3] = LodgeTown[lodge]
                                else:
                                    print("********FUCKED UP********" + loc)
                                    if loc.strip() in BadLocations:
                                        print("*!*!*!*!*!*!* BAD LOCATIONS *!*!*!*!*!*!*!*!")
                                        print(loc)
                                        full_list_calendar_entries[array_idx][3] = BadLocations[loc.strip()]
                                    else:
                                        location_exceptions.append(loc)
                                        full_list_calendar_entries[array_idx][3] = loc
                            else:
                                print("TRYING BAD LOC")
                                print(loc)
                                if loc.strip() in BadLocations:
                                    print("*!*!*!*!*!*!* BAD LOCATIONS *!*!*!*!*!*!*!*!")
                                    print(loc)
                                    full_list_calendar_entries[array_idx][3] = BadLocations[loc.strip()]
                                else:
                                    location_exceptions.append(loc)
                                    full_list_calendar_entries[array_idx][3] = loc
        array_idx += 1
        print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    print(full_list_calendar_entries)

def modify_date(table):
    time_array = [item[4] for item in table]
    print(time_array)
    curr_yr = datetime.datetime.today().year;

    array_idx = 0
    for time in time_array:
        print("-------------------------------------------")
        print(array_idx)
        print(full_list_calendar_entries[array_idx][0])
        print(time)
        time_str = time.strftime("%a %b. %-d")
        if time.year > curr_yr:
            print(time.year)
            time_str += ", "
            time_str += str(time.year)
        print(time_str)
        time_str += ", "
        if time.minute == 00:
            time_str += time.strftime("%-I ")
        else:
            time_str += time.strftime("%-I:%M ")
        if time.hour > 12:
            time_str += "p.m."
        else:
            time_str += "a.m."
        print(time_str)
        print(full_list_calendar_entries[array_idx][0])
        full_list_calendar_entries[array_idx][4] = time_str
        array_idx += 1

