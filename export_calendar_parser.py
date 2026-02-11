# This python file will parse the calendar dump Excel spreadsheet
# and populate data structures that can be used by export_doc() to
# create a Word document of events.
import datetime
import re
from typing import Any

import numpy
import logging

from openpyxl.reader.excel import load_workbook

from data_structs import Col, lodge_town, known_bad_locations, lodge_locations, location_exceptions, full_list_calendar_entries, has_dinner, \
    degrees

logger = logging.getLogger(__name__)

def export_calendar_data() -> None:
    """ This function will open the Excel workbook, choose the 'data' sheet,
    (ignoring the 'calc' sheet) and then export data and modify it as needed."""
    logger.debug("starting export_calendar_data")

    workbook = load_workbook('CalDumpFGW.xlsx')
#    workbook = load_workbook('FW_Test_1.xlsx')
    worksheet = workbook['data']

    # Find out how many rows are populated -- this is different than the max_rows.
    populated_rows = len([row for row in worksheet if any(cell.value is not None for cell in row)])
    logger.debug("populated rows: " + str(populated_rows))

    # initialize a list of lists which represents the spreadsheet, initialized to -1's,
    # to hold our data when transformed.
    # Also initialize a list to hold lodge locations when there is some doubt
    for i in range(populated_rows - 1):
        full_list_calendar_entries.insert(i, [-1, -1, -1, -1, -1])
        lodge_locations.insert(i, '')

    if not full_list_calendar_entries[populated_rows - 1]:
        full_list_calendar_entries.pop(populated_rows - 1)

    # Set the ending cell identifier...Since there are five columns, that will be
    # column "E" with the last populated row number (i.e., E65)
    ending_cell = 'E' + str(populated_rows)

    logger.debug("ending cell: " + str(ending_cell))

    # Row 1 is always a heading row, so build the table starting at row 2 (cell A2)
    table = numpy.array([[cell.value for cell in col] for col in worksheet['A2':ending_cell]])

    # For the five columns, modify the data into a format that is useable.
    # For lodge, put it into the lodge format according to the style guild:  {Name} No. {num}
    # For title and descr, specifically look for "degree" or "dinner" and drop the row index into a list for later
    # For location, this is complex -- try to find the lodge TOWN (only)
    # For date, modify from xx/xx/xxxx hh:mm:ss into "Sun, Feb 8, 8 p.m."
    modify_lodge(table)
    modify_event_title(table)
    modify_event_descr(table)
    modify_event_location(table)
    modify_date(table)

def modify_lodge(table: numpy.ndarray) -> None:
    """ This function creates a list of the lodge name from the first column in every list (row) in the
    table. The table is a list of lists of the spreadsheet values, and this list is all
    lodge names.  It then modifies those lodge names from "Lodge {num} {name} to
    {Name} Lodge No. {num} according to a regex
    This cell will be populated in every row."""
    lodge_array = [item[Col.LODGE.value] for item in table]

    # Since we will use it many times, compile the regex for the lodge name as
    # written in the cell (i.e., "Lodge 001 Hiram") and capture the values.
    r_lodge = re.compile(r'(Lodge) (\d{3}) ([a-zA-Z\-\'\. ]+)')

    # Loop through the list elements and if the regex matches, the modify the string
    # (i.e., from 'Lodge 001 Hiram' to 'Hiram Lodge No. 1')
    # This cell will be populated in every row
    array_idx = 0

    for lodge in lodge_array:
        result = r_lodge.match(lodge)
        if result:
             full_list_calendar_entries[array_idx][Col.LODGE.value] =\
                f"{result.group(3)} {result.group(1)} No. {result.group(2).lstrip('0')}"
        array_idx += 1

def modify_event_title(table: numpy.ndarray) -> None:
    """ Create a list of the values in the second column in every list (row) in the table.
    The table is a list of lists of the spreadsheet values and this list is all event descriptions.
    The aim here is to see if "degree" is in the title. If so, add the index to the list of degrees.
    This cell will be populated in every row."""
    event_array = [item[Col.TITLE.value] for item in table]

    r_degrees = re.compile(r'.*[Dd]egree.*')

    # Basically we're looking to see if there is a degree mentioned in the title,
    # otherwise we're not going to modify it.  If there's an instance of degree,
    # put the array index into the degree set -- we'll use that later to prepend
    # the line in the doc with "DEGREE" so that it can easily be identified.
    array_idx = 0
    for event in event_array:
        result = r_degrees.match(event)
        if result is not None:
            degrees.add(array_idx)
        full_list_calendar_entries[array_idx][Col.TITLE.value] = event
        array_idx += 1

    logger.debug(" Rows Containing Degrees: " + str(degrees))

def modify_event_descr(table: numpy.ndarray) -> None:
    """Create a list of the event descriptions.
    The aim here is to see if it contains "degree" or "dinner". If so, add the array index to
    the appropriate set.
    This cell may not be populated in every row so need to check that is it not empty
    before using it"""
    event_descr_array = [item[Col.DESCR.value] for item in table]

    r_degrees = re.compile(r'.*[Dd]egree.*')
    r_dinner = re.compile(r'.*[Dd]inner.*')

    # Similar to the event title, we're looking to see if they mention degree in here.
    # Also, this is usually where they mention 'dinner', so we will add the array index
    # to the dinner or degrees sets
    array_idx = 0
    for event_descr in event_descr_array:
        if event_descr:
            result = r_degrees.match(event_descr)
            if result:
                degrees.add(array_idx)
            result2 = r_dinner.search(event_descr)
            if result2:
                has_dinner.add(array_idx)
        full_list_calendar_entries[array_idx][Col.DESCR.value] = event_descr
        array_idx += 1

    logger.debug(" Rows Containing Degrees: " + str(degrees))
    logger.debug(" Rows Containing Has Dinner: " + str(has_dinner))

def modify_event_location(table:numpy.ndarray) -> None:
    """ Create and modify a list of event descriptions. The aim here is to find the
    town. This is the trickiest part of the entire export process.
    Use regex to match typical entries for event location to try to
    parse the town out of them.  If that doesn't work, check against known_bad_locations, or
    put a town in from the lodge-town dictionary.  We cannot just use the dictionary because sometimes a
    location is off-site from the lodge
    This cell is not always populated in every row so need to check that first"""
    event_loc_array = [item[Col.LOCATION.value] for item in table]
    array_idx = 0

    r_ct_usa = re.compile(r'.*[,]* (.*), (CT USA).*')
    r_ct_zip = re.compile(r'[\w ]*[,]? (.*), (CT \d{5})[,]*.*')
    r_lodge_name = re.compile(r'(Lodge) (\d{3}) ([a-zA-Z\-\'\. ]+)')
    r_just_ct = re.compile(r'.*, (.*), (CT)[,]* .*')
    r_just_town = re.compile(r'.*, (.*)')

    result = re.Match
    result2 = re.Match

    # We could just put the location in based on the lodge dictionary, but sometimes the event is not at the lodge.
    #
    # Look through all the locations and try to extract a town name.  This gets very tricky.  While there are
    # some locations that are easily parsed, some are more difficult.  We'll maintain two
    # lists -- one of lodge_locations (what we parse) and one of locations exceptions (an array index of what we
    # could not parse at all.)
    #
    # In all instances where there is a location field present, we'll put the
    # town in if we can find it into the doc, and the location field too -- the decision of which to use can
    # be made when editing the doc.
    logger.debug("location_exceptions before: "  + str(location_exceptions))

    for event_loc in event_loc_array:
        if not event_loc:
            # There was no location specified in the spreadsheet.  However, there is always
            # a lodge name.  Use the lodge name to grab the lodge location from the lodge_town dictionary
            full_list_calendar_entries[array_idx][Col.LOCATION.value] = lodge_town[table[array_idx][Col.LODGE.value]]
            array_idx += 1
            continue

        logger.debug("array_idx: " + str(array_idx))

        # Most of the cases here will populate full_list_calendar_entries with the event_loc.  Do that here.
        # If there is an exception later, it can be overwritten.
        full_list_calendar_entries[array_idx][Col.LOCATION.value] = event_loc

        # By now we know there is a location in the cell. Start by trying to pull the town out.
        # BUT, this might be a legit address for a non-lodge event (say pizza night at a pizza place).
        # Leave the location cell alone, but put the possible town into a separate location array
        #
        # Start by matching against the CT {zip} regex, if you get a match see if group(1) is just a town
        if (result := is_ct_zip(event_loc, r_ct_zip)) is not None:
            if (result2 := is_just_town(result.group(1), r_just_town)) is not None:
                lodge_locations[array_idx] = result2.group(1)
            else:
                # if not just a town, go with the first results into lodge_locations
                lodge_locations[array_idx] = result.group(1)
            array_idx += 1
            continue

        # no CT zip in the cell, so maybe it's just CT so set lodge_locations
        if (result2 := is_just_ct(event_loc, r_just_ct)) is not None:
            lodge_locations[array_idx] = result2.group(1)
            array_idx += 1
            continue

        # Not just CT, see if it is CT USA and set that in lodge_locations
        if (result2 := is_ct_usa(event_loc, r_ct_usa)) is not None:
            lodge_locations[array_idx] = result2.group(1)
            array_idx += 1
            continue

        # Haven't hit the usual patterns so at this point
        # just see if it is a lodge name -- if so, set the
        # cell to the value from the lodge town dictionary
        if fullmatch_lodge_name(event_loc, r_lodge_name):
            full_list_calendar_entries[array_idx][Col.LODGE.value] = lodge_town[event_loc]
            array_idx += 1
            continue

        # Did not fully match a lodge name in the event loc, so search for it.  If
        # the search finds it in the event_loc. construct the original lodge format
        # and see if it is in the  lodge town dictionary
        if (result2 := search_lodge_name(event_loc, r_lodge_name)) is not None:
            lodge = result2.group(1) + " " + result2.group(2) + " " + result2.group(3)
            logger.debug(" Search Found Lodge Name: |" + event_loc + " | " + lodge)
            try:
                if lodge in lodge_town.keys():
                    # found the lodge in the dictionary, so set the location in the cell
                    logger.debug("Lodge in lodge_town.keys() " + "|" + lodge + "|")
                    full_list_calendar_entries[array_idx][Col.LODGE.value] = lodge_town[lodge]
                    logger.debug("Put lodge into full list")
                else:
                    # didn't find it in the dictionary.  Check known_bad_locations
                    logger.debug("Lodge not in town keys")
                    if event_loc.strip() in known_bad_locations:
                        logger.debug("found in known_bad_locations, using that")
                        full_list_calendar_entries[array_idx][Col.LOCATION.value] = \
                            known_bad_locations[event_loc.strip()]
                    else:
                        logger.debug("Lodge not in town keys or known_bad_locations")
                        location_exceptions.append(event_loc)
                        logger.debug(" Location exception added: " + event_loc)
                        full_list_calendar_entries[array_idx][Col.LODGE.value] = event_loc

            except KeyError:
                # Trying the above caused an exception -- if it is in known_bad_locations,
                # put that in the list, otherwise put the the event loc in the exceptions list
                # and it can be sorted out when editing.
                logger.debug("Key Exception: " + lodge)
                if event_loc.strip() in known_bad_locations:
                    full_list_calendar_entries[array_idx][Col.LOCATION.value] = \
                        known_bad_locations[event_loc.strip()]
                else:
                    location_exceptions.append(event_loc)
                    logger.debug(" Location exception added: " + event_loc)
                    full_list_calendar_entries[array_idx][Col.LODGE.value] = event_loc

            array_idx += 1
            continue

        # If we got this far it's a bad location. If not already in known_bad_locations
        # add it
        if event_loc.strip() in known_bad_locations:
            # If the lodge is already in bad locations, put it in the cell
            full_list_calendar_entries[array_idx][Col.LOCATION.value] = \
                known_bad_locations[event_loc.strip()]
        else:
            # Not there, so put it in locations exceptions....
            location_exceptions.append(event_loc)
            array_idx += 1

    logger.debug("location_exceptions after: " + str(location_exceptions))

def modify_date(table: numpy.ndarray) -> None:
    time_array = [item[Col.DATE.value] for item in table]
    """ Create a list of the event times.
    These come in a mm/dd/yyyy hh:mm:ss and need to be changed to 
    something like Sun Feb 8, 8 p.m."""
    # Find the current year.  The style guide for our doc says we don't use the year
    # if it is in the current year, so grab this to compare later.
    curr_yr = datetime.datetime.today().year

    array_idx = 0
    for time in time_array:
        time_str = format_time(time, curr_yr)
        full_list_calendar_entries[array_idx][Col.DATE.value] = time_str
        array_idx += 1


def is_ct_zip(event_loc: str, regex: re.Pattern) -> re.Match:
    result = regex.match(event_loc)
    if result:
        return result
    else:
        return None
def is_just_town(event_loc: str, regex: re.Pattern) -> re.Match:
    result = regex.match(event_loc)
    if result:
        return result
    else:
        return None

def is_just_ct(event_loc: str, regex: re.Pattern) -> re.Match:
    result = regex.match(event_loc)
    if result is not None:
        return result
    else:
        return None

def is_ct_usa(event_loc: str, regex: re.Pattern) -> re.Match:
    result = regex.match(event_loc)
    if result:
        return result
    else:
        return None

def fullmatch_lodge_name(event_loc: str, regex: re.Pattern) -> re.Match:
    result = regex.fullmatch(event_loc)
    if result:
        return result
    else:
        return None

def search_lodge_name(event_loc: str, regex: re.Pattern) -> re.Match:
    result = regex.search(event_loc)
    if result:
        return result
    else:
        return None



def format_time(time: Any, curr_yr: int) -> str:
    """ Convert the time string into a different format"""
    time_str = time.strftime("%a %b. %-d")
    logger.debug(time_str)
    # If the year is greater than the current year, add it to the time_str so it prints.
    if time.year > curr_yr:
        time_str += ", "
        time_str += str(time.year)
    time_str += ", "

    # Style guide is to not print the :00 if it is 00 so do that.
    if time.minute == 00:
        time_str += time.strftime("%-I ")
    else:
        time_str += time.strftime("%-I:%M ")

    # We use a.m. and p.m. so make te appropriate choice.
    if time.hour > 12:
        time_str += "p.m."
    else:
        time_str += "a.m."

    return time_str
